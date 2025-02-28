from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

file_path = 'estoque.xlsx'
workbook = load_workbook(file_path)

# Correção do typo: workkbook -> workbook
sheet = workbook["Estoque"]

graph_sheet = workbook.create_sheet(title='Graficos')

max_row = sheet.max_row

col_nome_produto = 1
col_valor_fornecedor = 2
col_lucratividade = 3
col_quantidade = 4
col_preco_venda = 5
col_lucro_total = 6
col_valor_total = 7

# Verifica se a última linha é de totais e ajusta
if sheet.cell(row=max_row, column=col_nome_produto).value == "Totais Gerais":
    max_row -= 2

# Gráfico de Barras para Valor Total
bar_chart_valor_total = BarChart()
bar_chart_valor_total.title = "Valor Total em Estoque por Produto"
bar_chart_valor_total.y_axis.title = 'Valor Total (R$)'
bar_chart_valor_total.x_axis.title = 'Produto'

data = Reference(sheet, min_col=col_valor_total, min_row=1, max_row=max_row)
cats = Reference(sheet, min_col=col_nome_produto, min_row=2, max_row=max_row)

bar_chart_valor_total.add_data(data, titles_from_data=True)
bar_chart_valor_total.set_categories(cats)
bar_chart_valor_total.width = 30
bar_chart_valor_total.height = 15

graph_sheet.add_chart(bar_chart_valor_total, 'A1')

# Processamento da Lucratividade
produtos_lucratividades = []

for row in range(2, max_row + 1):
    nome_produto = sheet.cell(row=row, column=col_nome_produto).value
    lucratividade = sheet.cell(row=row, column=col_lucratividade).value
    
    if isinstance(lucratividade, (int, float)):
        produtos_lucratividades.append((nome_produto, lucratividade))

if produtos_lucratividades:
    # Ordena e pega top 5
    produtos_lucratividades.sort(key=lambda x: x[1], reverse=True)
    top_5_lucrativos = produtos_lucratividades[:5]

    # Cria sheet auxiliar
    aux_sheet = workbook.create_sheet(title="Auxiliar")
    aux_sheet.append(['Nome do Produto', 'Lucratividade (%)'])

    for nome, lucro in top_5_lucrativos:
        aux_sheet.append([nome, lucro])

    # Gráfico de Pizza
    pie_chart = PieChart()
    pie_chart.title = "Top 5 Produtos com Maior Lucratividade"
    
    data = Reference(aux_sheet, min_col=2, min_row=2, max_row=6)
    labels = Reference(aux_sheet, min_col=1, min_row=2, max_row=6)
    
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    
    # Configura rótulos
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showVal = True
    pie_chart.dataLabels.showPercent = True
    
    graph_sheet.add_chart(pie_chart, "A25")

    # Gráfico de Linhas para Quantidade (Exemplo corrigido)
    line_chart = LineChart()
    line_chart.title = "Quantidade em Estoque por Produto"
    
    # Pega dados da coluna de quantidade
    data = Reference(sheet, min_col=col_quantidade, min_row=1, max_row=max_row)
    cats = Reference(sheet, min_col=col_nome_produto, min_row=2, max_row=max_row)
    
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)
    
    graph_sheet.add_chart(line_chart, "A50")

# Salva fora do bloco condicional para garantir
workbook.save(file_path)